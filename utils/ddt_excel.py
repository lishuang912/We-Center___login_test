from openpyxl import load_workbook


class ParseExcel():
    def __init__(self, excelPath, sheetName):
        self.wb = load_workbook(excelPath)
        # self.wb.active 用于载入sheet
        self.sheet = self.wb[sheetName]
        self.maxRowNum = self.sheet.max_row

    def getDatasFromSheet(self):
        dataList = []
        for line in self.sheet:
            tmpList = []
            tmpList.append(line[0].value)
            tmpList.append(line[1].value)
            dataList.append(tmpList)
        return dataList[1:]


if __name__ == '__main__':
    excelPath = r'E:\selenium项目\by_shuang\selenium_test_shuang\data\yongli.xlsx'
    sheetName = 'login'
    pe = ParseExcel(excelPath, sheetName)
    for i in pe.getDatasFromSheet():
        print(i)
    print(pe.getDatasFromSheet())